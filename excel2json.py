#!/usr/bin/env python3
#usage - python excel2json.py <path to excel file>
import openpyxl
import json
import sys
import os

file = timetowait = sys.argv[1]
filename, extension = os.path.splitext(file);

wb = openpyxl.load_workbook(file)
sheetnames = wb.sheetnames
data = {}
for sheet in sheetnames:
    props = []
    data[sheet] = []
    sheet_obj = wb[sheet]
    rows = sheet_obj.rows
    for row in rows:
        rowdata = {}
        if len(props) is 0:
            for cell in row:
                props.append(cell.value)
        else:
            for index, cell in enumerate(row):
                rowdata[props[index]] = cell.value
            data[sheet].append(rowdata)

with open(filename + ".json", "w") as outfile:
    json.dump(data, outfile)
